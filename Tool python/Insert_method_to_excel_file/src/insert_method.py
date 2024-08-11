import os
from openpyxl import load_workbook

def manipulate_excel_file(directory, filename):
    """Thao tác với file Excel trong thư mục để kiểm tra sheet và chèn công thức."""
    
    file_path = os.path.join(directory, filename)

    # Kiểm tra xem file có tồn tại không
    if not os.path.isfile(file_path):
        print(f"File {file_path} không tồn tại.")
        return

    # Load workbook và kiểm tra sheet
    try:
        workbook = load_workbook(file_path, data_only=False)
        sheet_names = workbook.sheetnames
        
        if 'STD' not in sheet_names:
            print("Không có sheet STD.")
            return

        # Lấy sheet STD
        sheet = workbook['STD']

        # Thêm công thức vào cột T từ hàng 2 đến hết
        for row in range(2, sheet.max_row + 1):
            formula = f'="STD-"&D{row}&"-"&F{row}&"-"&H{row}'
            sheet[f'T{row}'] = formula

        # Lưu workbook sau khi thay đổi
        workbook.save(file_path)
        print(f"Đã cập nhật file: {file_path}")

    except Exception as e:
        print(f"Không thể thao tác với file {file_path}: {e}")

def insert_method_main(output_dir):
    manipulate_excel_file(output_dir)
